from openpyxl import Workbook
import random
import string

# Crée un nouveau classeur
wb = Workbook()

# Active la feuille par défaut
ws = wb.active

# Donne un nom à la feuille
ws.title = "MaFeuille"

# Écrit des données dans des cellules
#ws['A1'] = "Nom"
#ws['B1'] = "Âge"
lenght = random.randint(5, 10)

alphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]

for i in range(lenght):
    ws[alphabet[i].capitalize()+'1'] = alphabet[-i]

#ws.append(["Alice", 30])
#ws.append(["Bob", 25])
#ws.append(["Charlie", 35])
for i in range(random.randint(30, 50)):
    ws.append([str(i)] + [chr(random.randint(65, 90))for i in range(lenght)])

# Sauvegarde le fichier
wb.save("exemple.xlsx")
